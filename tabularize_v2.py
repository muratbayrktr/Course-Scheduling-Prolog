import re
from collections import defaultdict
import csv
import pickle
import sys
filename = sys.argv[1] if sys.argv[1] else "schedule.xlsx"
# Sample output from Clingo
with open("out.lp", "r") as f:  
    clingo_output = f.read()


# Read mapping.pickle dictionary obj.
with open("mapping.pickle", "rb") as f:
    mapping = pickle.load(f)

inverse_mapping = {v: k for k, v in mapping.items()}

# print(inverse_mapping)

# We will create an excel file with the following columns:
# Mon, Tue, Wed, Thu, Fri, Sat, Sun
# following rows will be the timeslots
# 08:30-09:30, 09:30-10:30, 10:30-11:30, 11:30-12:30, 12:30-13:30, 13:30-14:30, 14:30-15:30, 15:30-16:30, 16:30-17:30, 17:30-18:30, 18:30-19:30, 19:30-20:30
# We will fill the cells with the course codes
# We will use the mapping dictionary to map the course codes to the cells
# step1: creeate the excel file
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Schedule"
# columns:
columns = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
# rows:
rows = ["08:30-09:30", "09:30-10:30", "10:30-11:30", "11:30-12:30", "12:30-13:30", "13:30-14:30", "14:30-15:30", "15:30-16:30", "16:30-17:30", "17:30-18:30", "18:30-19:30", "19:30-20:30"]
# Write columns
for i, column in enumerate(columns):
    sheet.cell(row = 1, column = i+2).value = column

# Write rows
for i, row in enumerate(rows):
    sheet.cell(row = i+2, column = 1).value = row
# Save the workbook
wb.save(filename)


occupies_list = clingo_output.split()
# Create a dictionary
occupies_dict = defaultdict(list)
# Fill the dictionary
regex_pattern = r"occupies\((\d+),(\d+),(\w+),(\w+),(\d+),(\d+),(\d+)\)"

# Open excel file
wb = openpyxl.load_workbook(filename)
sheet = wb.active
# Fill the cells
for item in occupies_list:
    # print(item)
    # Extract the course code
    match = re.match(regex_pattern, item)
    if match is None:
        continue
    course_code = match.group(1)
    section = int(match.group(2))
    room = match.group(3)
    day, time = inverse_mapping[int(match.group(5))].split(" ")
    row_index = rows.index(time)
    excel_row_index = row_index + 2
    column_index = columns.index(day)
    excel_column_index = column_index + 2
    hours = int(match.group(7))
    teacher = match.group(4)

    print(course_code, section, room, day, time, hours, excel_row_index, excel_column_index)
    # Write to the excel file
    # Open the workbook
    # If the cell is empty, write the course code
    if sheet.cell(row = excel_row_index, column = excel_column_index).value is None:
        sheet.cell(row = excel_row_index, column = excel_column_index).value = course_code + "_" + str(section) + "_" + room + "_" + str(hours) + "_" + teacher
    else:
        # If the cell is not empty, append the course code
        sheet.cell(row = excel_row_index, column = excel_column_index).value += "\n" + course_code + "_" + str(section) + "_" + room + "_" + str(hours) + "_" + teacher
# Save the workbook
wb.save(filename)
wb.close()





