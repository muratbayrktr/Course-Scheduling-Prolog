import openpyxl
import sys


def sanitize_string(s):
    # Mapping of special Turkish characters to their English equivalents
    char_mapping = {
        'ğ': 'g',
        'Ğ': 'G',
        'ş': 's',
        'Ş': 'S',
        'ı': 'i',
        'İ': 'I',
        'ö': 'o',
        'Ö': 'O',
        'ü': 'u',
        'Ü': 'U',
        'ç': 'c',
        'Ç': 'C'
    }

    # Convert to lowercase
    sanitized = s.lower()

    # Replace special Turkish characters
    sanitized = ''.join(char_mapping.get(char, char) for char in sanitized)

    # Replace spaces with underscores
    sanitized = sanitized.replace(" ", "_")
    
    # Remove any non-alphanumeric characters except underscores
    sanitized = ''.join(char for char in sanitized if char.isalnum() or char == '_')
    
    return sanitized

def sanitize_time(s):
    sanitized = s.lower()
    sanitized = sanitized.replace(" ", "_")
    sanitized = sanitized.replace("-", "_")
    sanitized = sanitized.replace(":", "_")
    sanitized = ''.join(char for char in sanitized if char.isalnum() or char == '_' )
    return sanitized

class ExtractFacts:
    def __init__(self, workbook, sheet_name):
        self.workbook = workbook
        # Get the worksheet
        self.worksheet = self.workbook[sheet_name]       # Get the column names
        self.column_names = [cell_obj.value for cell_obj in self.worksheet[1]]
        # print(self.column_names)

    def process(self, from_row = 1, to_row = None, from_column = 1, to_column = None):
        # Get the max row count
        if to_row is None:
            max_row = self.worksheet.max_row
        else:
            max_row = to_row

        # Get the max column count
        if to_column is None:
            max_column = self.worksheet.max_column
        else:
            max_column = to_column

        # Iterate through the rows
        for i in range(from_row, max_row + 1):
            # Iterate through the columns
            for j in range(from_column, max_column + 1):
                # Get the cell object
                cell_obj = self.worksheet.cell(row = i, column = j)
                yield cell_obj
                
    def iterrows(self, from_row = 1, to_row = None, from_column = 1, to_column = None):
        # Get the max row count
        if to_row is None:
            max_row = self.worksheet.max_row
        else:
            max_row = to_row

        # Get the max column count
        if to_column is None:
            max_column = self.worksheet.max_column
        else:
            max_column = to_column


        # Iterate through the rows
        for i in range(from_row, max_row + 1):
            # Iterate through the columns
            temp = []
            for j in range(from_column, max_column + 1):
                # Get the cell object
                cell_obj = self.worksheet.cell(row = i, column = j)
                temp.append(cell_obj.value)
            yield temp
            # Yield row as a list

class Places(ExtractFacts):
    def __init__(self, workbook, sheet_name):
        super().__init__(workbook, sheet_name)

    def process(self, sheet_name):
        # A1->B7
        with open("places.lp", "w") as f:
            for value in super().iterrows(2, 7, 1, 2):
                place, max_capacity, *args = value
                fact = f"capacity({sanitize_string(place)},{max_capacity})."
                fact2 = f"room({sanitize_string(place)})."
                f.write(fact + "\n")
                f.write(fact2 + "\n")

        print("Created: ", "places.lp")



class Instructors(ExtractFacts):
    def __init__(self, workbook, sheet_name):
        super().__init__(workbook, sheet_name)

    def process(self, sheet_name):
     # A1->CI29
        mapping = {}
        day = 0
        with open("timeslots.lp", "w") as f:
            base = 1008
            slot = 0
            day_slot_count = 11
            for value in self.column_names[3:]:
                mapping[value] = base + slot
                fact = f"time_slot({base + slot})."
                f.write(fact + "\n")
                print(slot+base, slot, base, day_slot_count)
                if slot == day_slot_count:
                    base += 50

                    slot = 0
                else:
                    slot += 1
        # Save mapping dictionary to a file
        import pickle
        with open("mapping.pickle", "wb") as f:
            pickle.dump(mapping, f)
        
        with open("teaches.lp", "w") as f:
            with open("busy.lp", "w") as f2:
                for value in super().iterrows(2, 29):
                    instructor, course1, course2, *time_slot_busy = value
                    fact1 = f"teaches({sanitize_string(instructor)},{course1})."
                    fact2 = f"teaches({sanitize_string(instructor)},{course2})."
                    f.write(fact1 + "\n") if course1 else None
                    f.write(fact2 + "\n") if course2 else None

                    
                    for i, availability in enumerate(time_slot_busy):
                        if availability == "Yes":
                            fact = f"busy({sanitize_string(instructor)},{mapping[self.column_names[i+3]]})."
                            f2.write(fact + "\n")
        



        print("Created: ", "teaches.lp")
        print("Created: ", "busy.lp")
        print("Created: ", "timeslots.lp")

        # with open("durations.lp", "w") as f:
        #     for value in range(3, self.column_names.__len__()-3, 1):
        #         first, second, third = self.column_names[value], self.column_names[value+1], self.column_names[value+2]
        #         fact_2hours = f"duration({sanitize_time(first)},{sanitize_time(second)},2)."
        #         fact_3hours = f"duration({sanitize_time(first)},{sanitize_time(third)},3)."
        #         f.write(fact_2hours + "\n")
        #         f.write(fact_3hours + "\n")    
                
            
            

class Courses(ExtractFacts):
    def __init__(self, workbook, sheet_name):
        super().__init__(workbook, sheet_name)

    def process(self, sheet_name):
        # A1->G48
        with open("courses2.lp", "w") as f:
            section = 1
            last_course_name = ""
            for value in super().iterrows(2, 48):
                code, name, is_service, level, type, capacity, hours = value
                # fact1 = f"course({code},{sanitize_string(name)})."
                # fact2 = f"service({code})." if is_service == "Yes" else None
                # fact3 = f"level({code},{sanitize_string(level)})."
                # fact4 = f"type({code},{sanitize_string(type)})."
                # fact5 = f"capacity({code},{capacity})."
                # fact6 = f"hours({code},{hourse})."
                # f.write(fact1 + "\n")
                # f.write(fact2 + "\n") if fact2 else None
                # f.write(fact3 + "\n")
                # f.write(fact4 + "\n")
                # f.write(fact5 + "\n")
                # f.write(fact6 + "\n")
                if last_course_name == name:
                    section += 1
                else:
                    section = 1
                last_course_name = name
                fact = f"course({code},{section},{sanitize_string(name)},{sanitize_string(is_service)},{sanitize_string(level)},{sanitize_string(type)},{capacity},{hours})."
                f.write(fact + "\n")
                # f.write(fact_hours + "\n")

        print("Created: ", "courses2.lp")


        


if __name__ == "__main__":
    # Open the Excel file
    workbook = openpyxl.load_workbook('problem-data2.xlsx')

    # Get the sheet names
    sheet_names = workbook.sheetnames


    # Create a lambda function to instantiate the subclass by sheet_name
    sheet_class = lambda sheet_name: getattr(sys.modules[__name__], sheet_name.lower().capitalize())


    # Print the sheet names
    for sheet_name in sheet_names:
        print("Processing: ", sheet_name)
        # Find the subclass by sheet_name and instantiate 
        processor = sheet_class(sheet_name)(workbook,sheet_name)
        processor.process(sheet_name)



        
